import csv
import xlrd

from bs4 import BeautifulSoup
from celery import task
# Related third party imports.
from django.template.defaultfilters import slugify
from openpyxl import load_workbook

from agency.models import Agency, StatusChange
from main.models import Country, State, City
from main.tasks import send_mail

IATA_STATUS_DICT = {
    # Terminations And Closures
    'ter': 'T',
    'vrr': 'T',
    'clo': 'T',

    # Reinstatements
    'sdr': 'S',
    'dfw': 'S',
    'rtr': 'S',
    'rei': 'S',
    'rcr': 'S',

    # Default Information
    'dfe': 'D',

    # Reviews/Notices of Termination
    'not': 'R',
    'ntw': 'R',
    'rvw': 'R',

    # IRREGULARITIES AND ADMIN NONCOMPLIANCE
    'nfs': 'I',
    'rew': 'I',
    'lsp': 'I',
    'lsw': 'I',
}

ARC_STATUS_DICT = {
    'ACTIVE': 'A',
    'DEFAULTED': 'D',
    'REVOKED': 'R',
    'REINSTATED': 'S',
    'TERMINATED': 'T',
}

def extract_number(string):
    """Return the the agency_no in number from .
    extract_number('="01501846"') ->  01501846
    """
    # num = re.compile(r'="(?P<number>\d+)\s*"')
    # m = re.match(num, string)
    # vals = m.groupdict()
    return slugify(string).zfill(8)

@task(time_limit=999999, soft_time_limit=999999)
def process_agency_list_from_txt(filepath,country_name, is_async=False, to_email='assda@assda.com'):
    file_error = False
    with open(filepath) as f:
        html_doc = f.readlines()
        for line in html_doc:
            vals = line.split('|')
            if len(vals) > 1:
                if vals[7]:
                    if vals[7].strip() == 'Saint Pierre and Miquelon':
                        country, created_c = Country.objects.get_or_create(
                            name=country_name)
                    else:
                        country, created_c = Country.objects.get_or_create(
                            name=vals[7].strip())
                else:
                    country = None

                if vals[6]:
                    if vals[6].strip() == 'NS':
                        state, created_s = State.objects.get_or_create(
                            country=country, name='Nova Scotia')
                    else:
                        state, created_s = State.objects.get_or_create(
                            country=country, name=vals[6].strip())
                else:
                    state = None

                if vals[4]:
                    if state and country:
                        city, created_ci = City.objects.get_or_create(
                            country=country, state=state, name=vals[4].strip())
                    else:
                        city = None
                else:
                    city = None
                print({
                    'trade_name': vals[1].strip(),
                    'address1': vals[2].strip(),
                    'address2': vals[3].strip(),
                    'city': city,
                    'state': state,
                    # 'country': country,
                    'zip_code': vals[5].strip(),
                    'email': vals[10].lower().strip(),
                    'vat_number': vals[11].strip(),
                    'tel': vals[8].strip(),
                    # 'status_iata': 'A'
                })
                try:
                    agency, created = Agency.objects.update_or_create(agency_no=vals[0], country=country, defaults={
                        'trade_name': vals[1].strip(),
                        'address1': vals[2].strip(),
                        'address2': vals[3].strip(),
                        'city': city,
                        'state': state,
                        # 'country': country,
                        'zip_code': vals[5].strip(),
                        'email': vals[10].lower().strip(),
                        'vat_number': vals[11].strip(),
                        'tel': vals[8].strip(),
                        # 'status_iata': 'A'
                    })
                    old_status = agency.get_status_iata_display()
                    agency.status_iata = 'A'
                    agency.save()
                    new_status = agency.get_status_iata_display()
                    if old_status != new_status:
                        StatusChange.objects.create(old_status=old_status, new_status=new_status, reason='',
                                                    agency=agency)
                        if agency.status_iata == 'D':
                            context = {
                                'agency_no': agency.agency_no,
                                'agency_name': agency.trade_name,
                                'agency_status': 'defaulted'
                            }
                            send_mail("Agency status changed.", "email/status-default-email.html", context,
                                      ['abhilashanil2529@gmail.com'],
                                      from_email='assda@assda.com')
                except Exception as e:
                    print("eeeeeeeeeeeee",e)
                # print(agent, created)

    if is_async:
        context = {'user': to_email,
                   'file_name': filepath.split('/')[-1],
                   }
        send_mail('Agency list has been successfully uploaded', "email/agency-upload-email.html", context, [to_email], from_email='assda@assda.com')

    return file_error

@task(time_limit=999999, soft_time_limit=999999)
def process_agency_list_from_html(filepath,country_name ,is_async=False, to_email='assda@assda.com'):
    file_error = False
    with open(filepath) as f:
        html_doc = f.read()
        soup = BeautifulSoup(html_doc, 'html.parser')
        tds = soup.find_all("td", class_="line-content")
        for td in tds:
            vals = td.text.split('|')
            if len(vals) > 1:
                row = td.text
                if td.text.rfind('>'):
                    row = td.text[td.text.rfind('>') + 1:]

                vals = row.split('|')
                if vals[7]:
                    if vals[7].strip() == 'Saint Pierre and Miquelon':
                        country, created_c = Country.objects.get_or_create(
                            name=country_name)
                    else:
                        country, created_c = Country.objects.get_or_create(
                            name=vals[7].strip())
                else:
                    country = None

                if vals[6]:
                    if vals[6].strip() == 'NS':
                        state, created_s = State.objects.get_or_create(
                            country=country, name='Nova Scotia')
                    else:
                        state, created_s = State.objects.get_or_create(
                            country=country, name=vals[6].strip().title())
                else:
                    state = None

                if vals[4]:
                    if state and country:
                        city, created_ci = City.objects.get_or_create(
                            country=country, state=state, name=vals[4].strip().title())
                    else:
                        city = None
                else:
                    city = None

                agency, created = Agency.objects.update_or_create(agency_no=vals[0], country=country, defaults={
                    'trade_name': vals[1].strip(),
                    'address1': vals[2].strip(),
                    'address2': vals[3].strip(),
                    'city': city,
                    'state': state,
                    # 'country': country,
                    'zip_code': vals[5].strip(),
                    'email': vals[10].lower().strip(),
                    'vat_number': vals[11].strip(),
                    'tel': vals[8].strip(),
                    # 'status_iata':'A'
                })
                old_status = agency.get_status_iata_display()
                agency.status_iata = 'A'
                agency.save()
                new_status = agency.get_status_iata_display()
                if old_status != new_status:
                    StatusChange.objects.create(old_status=old_status, new_status=new_status, reason='',
                                                agency=agency)
                    if agency.status_iata == 'D':
                        context = {
                            'agency_no': agency.agency_no,
                            'agency_name': agency.trade_name,
                            'agency_status': 'defaulted'
                        }
                        send_mail("Agency status changed.", "email/status-default-email.html", context,
                                  [agency.email],
                                  from_email='Assda@assda.com')
        if len(tds) == 0:
            file_error = True
    if is_async:
        context = {'user': to_email,
                   'file_name': filepath.split('/')[-1],
                   }
        send_mail('Agency list has been successfully uploaded', "email/agency-upload-email.html", context, [to_email], from_email='assda@assda.com')

    return file_error

@task(time_limit=999999, soft_time_limit=999999)
def process_agency_list_from_csv(filepath, is_async=False, to_email='assda@assda.com'):
    file_error = False
    agency_data_entry = 0
    with open(filepath, encoding='utf-8', errors='replace') as csvfile:
        reader = csv.reader(csvfile, delimiter=',')
        next(reader, None)

        country, created_c = Country.objects.get_or_create(
            name="United States")

        for vals in reader:
            try:
                state = State.objects.filter(
                    country=country, abrev=vals[8].strip()).first()
                if not state:
                    state, created_s = State.objects.get_or_create(
                        country=country, abrev='OT', defaults={'name': 'Other'})
            except Exception as e:
                continue

            try:
                if state and country:
                    city, created_ci = City.objects.get_or_create(country=country, state=state, name=vals[7].strip())
            except Exception as e:
                city = None

            try:
                agency_no = extract_number(vals[1].strip())

                if agency_no:
                    try:
                        email = vals[58]
                    except IndexError as e:
                        email = vals[33]
                    # print(': ',email)
                    agency_data_entry = 1
                    agency, created = Agency.objects.update_or_create(agency_no=agency_no,
                                                                  country=country, defaults={
                            'trade_name': vals[26].strip(),
                            'address1': vals[5].strip(),
                            'address2': vals[6].strip(),
                            'city': city,
                            'state': state,
                            # 'country': country,
                            'zip_code': extract_number(vals[9].strip()),
                            'email': email.strip().lower(),
                            'vat_number': vals[11].strip(),
                            'tel': vals[10].strip(),
                            # 'home_agency': extract_number(vals[21].strip()),
                            # 'status':'A'
                        })
                    old_status = agency.get_status_display()
                    agency.status_iata = 'A'
                    agency.status = 'A'
                    agency.save()
                    new_status = agency.get_status_display()
                    if old_status != new_status:
                        StatusChange.objects.create(old_status=old_status, new_status=new_status, reason='',
                                                    agency=agency)
                        if agency.status == 'R':
                            context = {
                                'agency_no': agency.agency_no,
                                'agency_name': agency.trade_name,
                                'agency_status': 'revoked'
                            }
                            send_mail("Agency status changed.", "email/status-default-email.html", context,
                                      [agency.email],
                                      from_email='Assda@assda.com')

            except Exception as e:
                # avoid Tail, last row
                pass
        if not reader:
            file_error = True
        if agency_data_entry == 0:
            file_error = True
    if is_async:
        context = {'user': to_email,
                   'file_name': filepath.split('/')[-1],
                   }
        send_mail('Agency list has been successfully uploaded', "email/agency-upload-email.html", context, [to_email], from_email='assda@assda.com')

    return file_error


def process_bulletin(filepath, request=None):
    file_error = False
    agency_data_entry = 0
    file, extention = filepath.split('.')
    country = Country.objects.get(id=request.session.get('country'))
    ActList = ['REVOKE', 'REINSTATE', 'TERMINATE']

    # if extention.lower() in ['xlsx', 'xls']:
    wb = xlrd.open_workbook(filepath)

    for i in range(len(wb.sheet_names())):
        sht = wb.sheet_by_index(i)

        is_iata = {'Agency Code', 'Change Code'}.issubset(set(map(str.strip, sht.row_values(0))))
        is_arc = {'AGT_ACN', 'ACTION'}.issubset(set(map(str.strip, sht.row_values(0))))

        if is_arc and country.name == 'United States':
            for j in range(1,sht.nrows):
                values = sht.row_values(j)
                agency_no, status = values[1], values[8]

                if status in ActList:
                    status = status+'D'
                elif status == 'DEFAULT':
                    status = status+'ED'

                if agency_no and status:
                    # consider only some status changes
                    status_key = ARC_STATUS_DICT.get(status.upper(), None)
                    if status_key:
                        try:
                            agency_data_entry = 1
                            country, created_c = Country.objects.get_or_create(
                                name="United States")

                            try:
                                state = State.objects.get(country=country, name=values[7])
                            except Exception as e:
                                state, created_s = State.objects.get_or_create(
                                    country=country, abrev='OT', defaults={'name': 'Other'})

                            try:
                                if state and country:
                                    city, created_ci = City.objects.get_or_create(country=country, state=state,
                                                                                  name=values[6])
                            except Exception as e:
                                city = None

                            agency_no = extract_number(int(agency_no))

                            agency, created = Agency.objects.get_or_create(agency_no=agency_no, country=country,
                                                                           defaults={
                                                                               'trade_name': values[5],
                                                                               # 'address1': values[16],
                                                                               'city': city,
                                                                               'state': state,
                                                                               # 'country': country,
                                                                               # 'zip_code': values[17],
                                                                               # 'email': values[20],
                                                                               # 'vat_number': values[21],
                                                                               # 'tel': values[18],
                                                                           }
                                                                           )

                            old_status = agency.get_status_display()

                            # agency.status_iata = status_key
                            agency.status = status_key
                            agency.save()

                            new_status = agency.get_status_display()

                            if old_status != new_status:
                                StatusChange.objects.create(old_status=old_status, new_status=new_status,
                                                            reason=status,
                                                            agency=agency)
                                if agency.status == 'R':
                                    context = {
                                        'agency_no': agency.agency_no,
                                        'agency_name': agency.trade_name,
                                        'agency_status': 'revoked'
                                    }
                                    to = agency.sales_owner.email if agency.sales_owner else ''
                                    send_mail("Agency status changed.", "email/status-default-email.html", context,
                                              [to],
                                              from_email='Assda@assda.com')
                        except Exception as e:
                            pass
        elif is_iata :
            for j in range(1, sht.nrows):
                values = sht.row_values(j)
                agency_no, status = values[2], values[1]

                if agency_no and status:
                    # consider only some status changes
                    status_key = IATA_STATUS_DICT.get(slugify(status[:3]), None)
                    if status_key:
                        try:
                            agency_data_entry = 1

                            try:
                                state = State.objects.get(country=country, name=values[15])
                            except Exception as e:
                                state, created_s = State.objects.get_or_create(
                                    country=country, abrev='OT', defaults={'name': 'Other'})

                            try:
                                if state and country:
                                    city, created_ci = City.objects.get_or_create(country=country, state=state,
                                                                                  name=values[14])
                            except Exception as e:
                                city = None

                            agency_no = extract_number(int(agency_no))

                            agency, created = Agency.objects.get_or_create(agency_no=agency_no, country=country,
                                                                           defaults={
                                                                               'trade_name': values[3],
                                                                               'address1': values[16],
                                                                               'city': city,
                                                                               'state': state,
                                                                               # 'country': country,
                                                                               'zip_code': values[17],
                                                                               'email': values[20],
                                                                               'vat_number': values[21],
                                                                               'tel': values[18],
                                                                           }
                                                                           )
                            old_status = agency.get_status_iata_display()

                            agency.status_iata = status_key

                            agency.save()
                            new_status = agency.get_status_iata_display()

                            if old_status != new_status:
                                StatusChange.objects.create(old_status=old_status, new_status=new_status,
                                                            reason=status,
                                                            agency=agency)
                                if agency.status_iata == 'D':
                                    context = {
                                        'agency_no': agency.agency_no,
                                        'agency_name': agency.trade_name,
                                        'agency_status': 'defaulted'
                                    }
                                    to = agency.sales_owner.email if agency.sales_owner else ''
                                    send_mail("Agency status changed.", "email/status-default-email.html", context,
                                              [to],
                                              from_email='Assda@assda.com')
                        except Exception as e:
                            pass

        else:
            file_error = True

    # else:
    #     wb = load_workbook(filepath)
    #     wb = wb.worksheets
    #
    #     for sheet in wb:
    #         # iterating over the rows and
    #         # getting value from each cell in row
    #         reader = sheet.iter_rows()
    #         header = next(reader, None)
    #         # header_values = {*map(lambda x: x.value, header)}
    #
    #         # IATA Bulletin file check
    #         is_iata = {'Agency Code', 'Change Code'}.issubset({*map(lambda x: x.value, header)})
    #         is_arc = {'AGT_ACN', 'ACTION'}.issubset({*map(lambda x: x.value, header)})
    #         if is_iata and country.name == 'Canada':
    #             for row in reader:
    #                 values = [cell.value for cell in row]
    #                 agency_no, status = values[2], values[1]
    #                 if agency_no and status:
    #                     # consider only some status changes
    #                     status_key = IATA_STATUS_DICT.get(slugify(status[:3]), None)
    #                     if status_key:
    #                         try:
    #                             country, created_c = Country.objects.get_or_create(
    #                                 name="Canada")
    #
    #                             try:
    #                                 state = State.objects.get(country=country, name=values[15])
    #                             except Exception as e:
    #                                 state, created_s = State.objects.get_or_create(
    #                                     country=country, abrev='OT', defaults={'name': 'Other'})
    #
    #                             try:
    #                                 if state and country:
    #                                     city, created_ci = City.objects.get_or_create(country=country, state=state,
    #                                                                                   name=values[14])
    #                             except Exception as e:
    #                                 city = None
    #
    #                             agency_no = extract_number(agency_no)
    #                             agency, created = Agency.objects.get_or_create(agency_no=agency_no[:7], country=country,
    #                                                                            defaults={
    #                                                                                'trade_name': values[3],
    #                                                                                'address1': values[16],
    #                                                                                'city': city,
    #                                                                                'state': state,
    #                                                                                # 'country': country,
    #                                                                                'zip_code': values[17],
    #                                                                                'email': values[20],
    #                                                                                'vat_number': values[21],
    #                                                                                'tel': values[18],
    #                                                                            }
    #                                                                            )
    #                             old_status = agency.get_status_iata_display()
    #                             agency.status_iata = status_key
    #                             agency.save()
    #                             new_status = agency.get_status_iata_display()
    #                             if old_status != new_status:
    #                                 StatusChange.objects.create(old_status=old_status, new_status=new_status, reason=status,
    #                                                             agency=agency)
    #                                 if agency.status_iata == 'D':
    #                                     context = {
    #                                         'agency_no': agency.agency_no,
    #                                         'agency_name': agency.trade_name
    #                                     }
    #                                     to = agency.sales_owner.email if agency.sales_owner else ''
    #                                     send_mail("Agency status changed.", "email/status-default-email.html", context,
    #                                               [to],
    #                                               from_email='Assda@assda.com')
    #                         except Exception as e:
    #                             pass
    #         elif is_arc and country.name == 'United States':
    #             for row in reader:
    #                 values = [cell.value for cell in row]
    #                 agency_no, status = values[2], values[1]
    #                 if agency_no and status:
    #                     # consider only some status changes
    #                     status_key = ARC_STATUS_DICT.get(status, None)
    #                     if status_key:
    #
    #                         try:
    #                             country, created_c = Country.objects.get_or_create(
    #                                 name="United States")
    #
    #                             try:
    #                                 state = State.objects.get(country=country, name=values[7])
    #                             except Exception as e:
    #                                 state, created_s = State.objects.get_or_create(
    #                                     country=country, abrev='OT', defaults={'name': 'Other'})
    #
    #                             try:
    #                                 if state and country:
    #                                     city, created_ci = City.objects.get_or_create(country=country, state=state,
    #                                                                                   name=values[6])
    #                             except Exception as e:
    #                                 city = None
    #
    #                             agency_no = extract_number(agency_no)
    #                             agency, created = Agency.objects.get_or_create(agency_no=agency_no[:7], country=country,
    #                                                                            defaults={
    #                                                                                'trade_name': values[5],
    #                                                                                # 'address1': values[16],
    #                                                                                'city': city,
    #                                                                                'state': state,
    #                                                                                # 'country': country,
    #                                                                                # 'zip_code': values[17],
    #                                                                                # 'email': values[20],
    #                                                                                # 'vat_number': values[21],
    #                                                                                # 'tel': values[18],
    #                                                                            }
    #                                                                            )
    #
    #                             old_status = agency.get_status_display()
    #                             # agency.status_iata = status_key
    #                             agency.status = status_key
    #                             agency.save()
    #
    #                             new_status = agency.get_status_display()
    #                             if old_status != new_status:
    #                                 StatusChange.objects.create(old_status=old_status, new_status=new_status,
    #                                                             reason=status,
    #                                                             agency=agency)
    #                                 if agency.status.upper() == 'REVOKED' or agency.status.upper() == 'REVOKE':
    #                                     context = {
    #                                         'agency_no': agency.agency_no,
    #                                         'agency_name': agency.trade_name
    #                                     }
    #                                     to = agency.sales_owner.email if agency.sales_owner else ''
    #                                     send_mail("Agency status changed.", "email/status-default-email.html", context,
    #                                               [to],
    #                                               from_email='Assda@assda.com')
    #                         except Exception as e:
    #                             pass
    #         else:
    #             file_error = True
        if agency_data_entry == 0:
            file_error = True

    return file_error
